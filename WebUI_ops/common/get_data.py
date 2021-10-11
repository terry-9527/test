import configparser
import yaml, os
from openpyxl import load_workbook
from WebUI_ops.common.getpathInfo import getpathInfo
import pandas


class GetData():

    # 读取通用配置文件
    def getConfigData(self, section=None, option=None, file=None):
        rootpath = getpathInfo()
        if not file:
            configpath = os.path.join(rootpath, 'config', 'config.ini')
        else:
            configpath = os.path.join(rootpath, 'config', file)
        config = configparser.RawConfigParser()
        config.read(configpath, encoding='utf-8')
        self.result = config.get(section, option)
        return self.result

    def getYamlData(self, file):
        root_dir = getpathInfo()
        file_path = os.path.join(root_dir, 'testdata', file)
        with open(file_path, 'r', encoding='utf-8') as f:
            yaml_data = yaml.load(f, Loader=yaml.FullLoader) # 读取yaml文件内容，返回dict数据
            case_data = []
            for case in yaml_data.values():
                case_data.append(case)
            f.close()
        return case_data

    def getExcel(self, sheetname, file):
        root_path = getpathInfo()
        file_path = os.path.join(root_path, "testdata", file)
        workbook = load_workbook(file_path)
        sheet = workbook[sheetname]  # 执行使用哪个工作表,根据传进来的表名称决定读取对应的数据
        rows = sheet.rows  # 取出所有行的数据
        cases = []
        # 遍历每一行的数据
        for row in rows:
            list1 = []
            if row[len(row) - 2].value == "True":  # 判断该条用例是否需要执行,execute最后一列的值为True则执行，进行数据读取
                for col in row:
                    if not col.value:  # 处理空单元格，直接添加None
                        list1.append(col.value)
                    elif col.value.startswith('{') and col.value.endswith('}'):
                        list1.append(eval(col.value))  # 获取当前行每一个单元格，单元格的内容需要用value，把str类型转化为dict类型
                    else:
                        list1.append(col.value)
                cases.append(list1)
        return cases

    # Excel写入数据
    def writeExcel(self, file, case_id=None, testresult=None):
        root_path = getpathInfo()
        file_path = os.path.join(root_path, "testdata", file)
        workbook = load_workbook(file_path)
        sheetname = workbook.sheetnames  # 获取工作表名称
        sheet = workbook[sheetname[0]]  # 执行使用哪个工作表
        cols = sheet.columns
        col = [col for col in cols]
        list1 = []
        for row in col[0]:
            list1.append(row.value)
        i = 0
        for id in list1:
            if case_id == id:
                i += 1
                break
            i += 1
        # print(i)
        sheet.cell(i, sheet.max_column).value = testresult
        # print(sheet.cell(i, sheet.max_column).value)
        workbook.save(file_path)


if __name__ == '__main__':
    data = GetData()
    # result = data.getYamlData('test.yaml')
    # print(result)
    # cases = data.getExcel('logincase.xlsx')
    cases = data.getExcel('adduser', 'userinfo.xlsx')
    print(cases)

