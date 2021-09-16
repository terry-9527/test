
import openpyxl

def get_excel_data(path):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook["Sheet1"]
    rows = sheet.rows
    # print(rows)
    list_data=[]
    for row in rows:
        # print(row)
        if row[0].value == "ID":
            continue
        list1 = []
        for col in row:
            list1.append(col.value)
        list_data.append(list1)
    # print(list_data)
    return list_data

# get_excel_data("./testpost.xlsx")
#方式二
# def get_excel_data(path):
#     workbook = openpyxl.load_workbook(path)
#     sheet = workbook["Sheet1"]
#     max_row = sheet.max_row
#     list_data = []
#     for i in range(2,max_row+1):
#         data=dict(
#             id = sheet.cell(row=i,column=1).value,
#             url = sheet.cell(row=i,column=2).value,
#             data = sheet.cell(row=i,column=3).value,
#             code = sheet.cell(row=i,column=4).value,
#             excepted = sheet.cell(row=i,column=5).value
#         )
#         # print(data["id"],data["url"],data["data"],data["code"],data["excepted"] )
#         # print(data)
#         list_data.append(data)
#     print(list_data)
#     return list_data
# get_excel_data("./testpost.xlsx")

# 方法三
# def read_excel():
#     wb = openpyxl.load_workbook("./testpost.xlsx")
#     ws = wb.active
#     testcase_list = []
#     header_list = []
#     for row in range(1, ws.max_row + 1):
#         one_row_data = {}   #每一行数据信息
#         # print(one_cell_row)
#         for col in range(1, ws.max_column + 1):
#             one_cell_value = ws.cell(row,col).value #获取每个单元格的值
#             if row == 1:
#                 header_list.append(one_cell_value)  #判断如果是第一行，取出每一个单元格的值，保存到header_list列表中
#                 # print(header_list)
#             else:                                   #如果不是第一行，则把该行的每一个单元格的内容保存到字典当中
#                 key = header_list[col-1]
#                 one_row_data[key] = one_cell_value
#         if row != 1:
#             testcase_list.append(one_row_data)  #除去第一行，每一行的数据保存到列表当中
#     # print(testcase_list)
#     return testcase_list    #返回整个Excel表格中读取到的所有数据