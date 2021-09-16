import os
import shutil
import time
from multiprocessing import Process
import threading
import openpyxl
file = openpyxl.load_workbook('./test.xlsx')
print(file.sheetnames)
print(file.active)
sheet = file["64G"]
print(sheet)
print(sheet.max_row)
print(sheet.max_column)
print(sheet["A1"])
print(sheet.cell(row=1,column=1).value)
print(sheet['1'])
for i in sheet['1']:
    print(i.value)


# lock1 = threading.Lock()
# lock2 = threading.Lock()
#
# def func1():
#     lock1.acquire(timeout=2)
#     print('------func1-------->lock1')
#     time.sleep(3)
#     lock2.acquire(timeout=2)
#     print('------func1-------->lock2')
#
#     lock2.release()
#     lock1.release()
#
# def func2():
#     lock2.acquire(timeout=2)
#     print('------func2-------->lock2')
#     time.sleep(3)
#     lock1.acquire(timeout=2)
#     print('------func2-------->lock1')
#
#     lock1.release()
#     lock2.release()
#
# if __name__ == '__main__':
#     t1 = threading.Thread(target=func1,name='Thread_1')
#     t2 = threading.Thread(target=func2,name='Thread_2')
#     t1.start()
#     t2.start()
#     t1.join()
#     t2.join()


# from multiprocessing import Manager,Pool
# import time
#
# def write(q):
#     for i in ["A","B","C","D","E"]:
#         print("向队列中添加%s"%i)
#         q.put(i)
#         # time.sleep(1)
#
# def read(q):
#     while not q.empty():
#         print("从队列中取出的值是%s"%q.get())
#         time.sleep(2)
#
# if __name__ == '__main__':
#     q = Manager().Queue()
#     pool = Pool()
#     pool.apply_async(write,args=(q,))
#     pool.apply_async(read,args=(q,))
#     # pool.apply(write,args=(q,))
#     # pool.apply(read,args=(q,))
#
#     pool.close()
#     pool.join()
#
#     print("数据通信完毕")


# from multiprocessing import Pool
# import os, time, random
#
#
# def worker(msg):
#     print("%s开始执行,进程号为%d" % (msg, os.getpid()))
#     time.sleep(1)
#     print
#     "%s执行完毕" % (msg)
#
#
# if __name__ == '__main__':
#     po = Pool(3)  # 定义一个进程池，最大进程数3
#     for i in range(10):
#         # Pool.apply_async(要调用的目标,(传递给目标的参数元祖,))
#         # 每次循环将会用空闲出来的子进程去调用目标
#         po.apply_async(worker, (i,))
#
#     print("----start----")
#     po.close()  # 关闭进程池，关闭后po不再接收新的请求
#     po.join()  # 等待po中所有子进程执行完成，必须放在close语句之后
#     print("-----end-----")
