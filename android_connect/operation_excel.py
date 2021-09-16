import openpyxl
import configparser

# conf = configparser.ConfigParser()
#
# data = conf.read('./config.ini')
# # print(data)
# sections = conf.sections()
# print(sections)
#
# for section in sections:
#     options = conf.options(section)
#     # print(section)
#
# print(conf.get('logging','server'))
# print(conf.get('mysql','host'))
# print(conf.get('mysql','port'))
# a=conf.getboolean('mysql','open')
# if a:
#     print('xxxxxxx',a)
# x = lambda n:n**2
# print(x(2))
# for i in range(1,5):
#     print(x(i))



def read_excel():
    wb = openpyxl.load_workbook("./testpost.xlsx")
    ws = wb.active
    testcase_list = []
    header_list = []
    for row in range(1, ws.max_row + 1):
        one_row_data = {}   #每一行数据信息
        # print(one_cell_row)
        for col in range(1, ws.max_column + 1):
            one_cell_value = ws.cell(row,col).value #获取每个单元格的值
            if row == 1:
                header_list.append(one_cell_value)  #判断如果是第一行，取出每一个单元格的值，保存到header_list列表中
                # print(header_list)
            else:                                   #如果不是第一行，则把该行的每一个单元格的内容保存到字典当中
                key = header_list[col-1]
                one_row_data[key] = one_cell_value
        if row != 1:
            testcase_list.append(one_row_data)  #除去第一行，每一行的数据保存到列表当中
    print(testcase_list)
    return testcase_list    #返回整个Excel表格中读取到的所有数据

if __name__ == '__main__':
    read_excel()
    # print(ws.title)
    # ws['A1']="ID"
    # wb.save("./testpost.xlsx")
    # print(ws['A1'].value)
    # rows = ws.iter_rows()
    # print(rows)
    # print(ws.cell(row=1,column=2).value)
    # print(ws['B1'].value)
# for i in rows:
#     for col in i:
#         print(col.value)
# cols = ws['B:C']
# print(cols)
# rows1 = ws['2:4']
# print(rows1)
# a = ws.iter_rows(min_row=2,max_row=3,min_col=2,max_col=3)
# for i in a:
#     print(i)
#     for x in i:
#         print(x.value)
# ws.append([111,'http://arsyuntest.com','{"postcode":"215001","key":"7518fef83f201dbeeb882131e0b7503k","type":"xml"}'])
# ws.delete_rows(5,7)
# ws['E1']='excepted'
# print(ws.max_row)
# print(ws.max_column)
# print(ws['E1'].value)
# wb.save('./testpost.xlsx')
